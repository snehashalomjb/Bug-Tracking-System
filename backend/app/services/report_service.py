from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.models.models import Bug
from app.repositories.bug_repository import BugRepository

class ReportService:
    @staticmethod
    def get_filtered_bugs(
        db: Session,
        status: Optional[str] = None,
        priority: Optional[str] = None,
        severity: Optional[str] = None,
        project_id: Optional[int] = None,
        assignee_id: Optional[int] = None
    ) -> List[Bug]:
        """Fetch bugs using filters without pagination limits for reporting exports."""
        # Query up to 1000 bugs for reports
        items, _ = BugRepository.get_all_bugs(
            db,
            skip=0,
            limit=1000,
            status=status,
            priority=priority,
            severity=severity,
            project_id=project_id,
            assignee_id=assignee_id,
            sort_by="created_at",
            sort_desc=True
        )
        return items

    @staticmethod
    def generate_excel_report(bugs: List[Bug]) -> bytes:
        """Create a styled Excel spreadsheet byte array using openpyxl."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bugs Report"

        # Apply grid lines visibility
        ws.views.sheetView[0].showGridLines = True

        # Styles
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Header list
        headers = [
            "Bug ID", "Title", "Project", "Module", 
            "Priority", "Severity", "Status", 
            "Reporter", "Assignee", "Due Date", "Created Date"
        ]

        ws.append(headers)
        
        # Style headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
        ws.row_dimensions[1].height = 28

        # Append data
        for row_idx, bug in enumerate(bugs, start=2):
            ws.append([
                bug.id,
                bug.title,
                bug.project.name if bug.project else "N/A",
                bug.module_name or "N/A",
                bug.priority,
                bug.severity,
                bug.status,
                bug.reporter.full_name if bug.reporter else "N/A",
                bug.assignee.full_name if bug.assignee else "Unassigned",
                bug.due_date.strftime("%Y-%m-%d") if bug.due_date else "N/A",
                bug.created_at.strftime("%Y-%m-%d")
            ])
            
            # Format and border cell cells
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border_thin
                if col_idx in [1, 5, 6, 7, 10, 11]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        # Autofilter and adjust column widths
        ws.auto_filter.ref = f"A1:K{len(bugs) + 1}"
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

        file_stream = BytesIO()
        wb.save(file_stream)
        return file_stream.getvalue()

    @staticmethod
    def generate_pdf_report(bugs: List[Bug], title_text: str) -> bytes:
        """Create a premium styled PDF report using reportlab doc template."""
        buffer = BytesIO()
        # Landscape/Letter page size template with 0.5 in margins
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter, 
            rightMargin=36, 
            leftMargin=36, 
            topMargin=36, 
            bottomMargin=36
        )
        story = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            leading=24,
            textColor=colors.HexColor('#1A365D'),
            spaceAfter=6
        )
        meta_style = ParagraphStyle(
            'ReportMetadata',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor('#4A5568'),
            spaceAfter=15
        )

        story.append(Paragraph(title_text, title_style))
        story.append(Paragraph(f"Generated on: {date.today().strftime('%B %d, %Y')} | Total Records: {len(bugs)}", meta_style))
        story.append(Spacer(1, 5))

        # PDF Table headers
        data = [["ID", "Title", "Project", "Priority", "Status", "Assignee", "Created"]]
        for bug in bugs:
            # Shorten fields to fit width
            title = bug.title[:38] + "..." if len(bug.title) > 38 else bug.title
            project = bug.project.name[:14] if bug.project else "N/A"
            assignee = bug.assignee.full_name[:14] if bug.assignee else "Unassigned"
            
            data.append([
                str(bug.id),
                title,
                project,
                bug.priority,
                bug.status,
                assignee,
                bug.created_at.strftime("%Y-%m-%d")
            ])

        # Table width constraints (Page width 612, printable width 540)
        # We can map widths to sum up to 540
        t = Table(data, colWidths=[25, 185, 75, 55, 60, 80, 60])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Center IDs
            ('ALIGN', (3, 0), (4, -1), 'CENTER'),  # Center Priority/Status
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Alternating background colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')])
        ]))

        story.append(t)
        doc.build(story)
        return buffer.getvalue()
